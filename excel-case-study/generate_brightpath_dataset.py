from __future__ import annotations

from collections import Counter
from datetime import date, datetime, timedelta
import random
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


BASE_DIR = Path(__file__).resolve().parent
CLEAN_FILE = BASE_DIR / "employee_records_clean.xlsx"
RAW_FILE = BASE_DIR / "employee_records_raw.xlsx"
README_FILE = BASE_DIR / "README.md"

SEED = 20260707
TOTAL_RECORDS = 3500

DEPARTMENTS = [
    "Accounting",
    "Administration",
    "Customer Service",
    "Finance",
    "Human Resources",
    "Information Technology",
    "Marketing",
    "Operations",
    "Quality Assurance",
    "Sales",
]

POSITIONS = [
    "Data Encoder",
    "Administrative Assistant",
    "HR Assistant",
    "Customer Support Representative",
    "Team Leader",
    "Operations Associate",
    "Quality Analyst",
    "Payroll Assistant",
    "Office Assistant",
    "Recruitment Associate",
    "Software Support Specialist",
]

FIRST_NAMES = [
    "John", "Maria", "Carlo", "Angela", "Joshua", "Patricia", "Mark", "Jayson", "Raven",
    "Nicole", "Paolo", "Sabrina", "Leah", "Miguel", "Kevin", "Alyssa", "Erika", "Noel",
    "Janelle", "Kristine", "Rochelle", "Paula", "Ivan", "Janice", "Rey", "Ella", "Adrian",
    "Monica", "Gabriel", "Trisha", "Dylan", "Hazel", "Arvin", "Samantha", "Ruth", "Christian",
    "Bea", "Jerome", "Angelica", "Floyd", "Clarisse", "Neil", "Mika", "Dexter", "Lara",
    "Enzo", "Catherine", "Martin", "Shane", "Bianca", "Jared",
]

LAST_NAMES = [
    "Santos", "Reyes", "Cruz", "Garcia", "Mendoza", "Dela Cruz", "Ramos", "Flores", "Lopez",
    "Bautista", "Torres", "Navarro", "Soriano", "Castillo", "Villanueva", "Gutierrez", "Diaz",
    "Hernandez", "Santiago", "Rivera", "Pascual", "Domingo", "Aquino", "Mercado", "Ortega",
    "Padilla", "Rizal", "Ocampo", "Marquez", "Fernandez", "Salazar", "Lacson", "De Guzman",
    "Valdez", "Salvador", "Alvarez", "Dimaculangan", "Molina", "Cabrera", "Cortez",
]

CITIES_BY_PROVINCE = {
    "Quezon": ["Lucena", "Tayabas"],
    "Batangas": ["Lipa", "Batangas City", "Tanauan", "Sto. Tomas"],
    "Laguna": ["Calamba", "San Pablo", "Santa Rosa", "Cabuyao"],
    "Rizal": ["Cainta", "Antipolo", "Taytay"],
    "Cavite": ["Dasmarinas", "Imus", "Bacoor"],
}

POSITIONS_BY_DEPT = {
    "Accounting": ["Data Encoder", "Payroll Assistant"],
    "Administration": ["Administrative Assistant", "Office Assistant"],
    "Customer Service": ["Customer Support Representative", "Team Leader"],
    "Finance": ["Payroll Assistant", "Data Encoder"],
    "Human Resources": ["HR Assistant", "Recruitment Associate"],
    "Information Technology": ["Software Support Specialist", "Data Encoder"],
    "Marketing": ["Office Assistant", "Data Encoder"],
    "Operations": ["Operations Associate", "Team Leader", "Office Assistant"],
    "Quality Assurance": ["Quality Analyst", "Team Leader"],
    "Sales": ["Office Assistant", "Customer Support Representative", "Team Leader"],
}

STATUS_WEIGHTS = [
    ("Active", 0.78),
    ("Probationary", 0.1),
    ("Inactive", 0.06),
    ("Resigned", 0.06),
]

POSITION_SALARY = {
    "Data Encoder": (25000, 32000),
    "Administrative Assistant": (26000, 34000),
    "HR Assistant": (28000, 36000),
    "Customer Support Representative": (26000, 38000),
    "Team Leader": (45000, 65000),
    "Operations Associate": (28000, 40000),
    "Quality Analyst": (32000, 48000),
    "Payroll Assistant": (30000, 42000),
    "Office Assistant": (25000, 33000),
    "Recruitment Associate": (30000, 42000),
    "Software Support Specialist": (45000, 80000),
}


def weighted_choice(rng: random.Random, items: list[tuple[str, float]]) -> str:
    total = sum(weight for _, weight in items)
    threshold = rng.random() * total
    running = 0.0
    for value, weight in items:
        running += weight
        if threshold <= running:
            return value
    return items[-1][0]


def build_first_names() -> list[str]:
    names: list[str] = []
    while len(names) < TOTAL_RECORDS:
        names.extend(FIRST_NAMES)
    return names[:TOTAL_RECORDS]


def generate_clean_records() -> list[dict[str, object]]:
    rng = random.Random(SEED)
    first_names = build_first_names()

    dept_weights = [
        ("Customer Service", 0.23),
        ("Operations", 0.21),
        ("Sales", 0.11),
        ("Quality Assurance", 0.1),
        ("Administration", 0.08),
        ("Human Resources", 0.07),
        ("Accounting", 0.06),
        ("Finance", 0.06),
        ("Marketing", 0.04),
        ("Information Technology", 0.04),
    ]

    province_weights = [
        ("Laguna", 0.34),
        ("Batangas", 0.28),
        ("Quezon", 0.18),
        ("Cavite", 0.12),
        ("Rizal", 0.08),
    ]

    recent_start = date(2022, 1, 1)
    recent_end = date(2025, 12, 31)
    older_start = date(2018, 1, 1)
    older_end = date(2021, 12, 31)

    records: list[dict[str, object]] = []
    for index in range(TOTAL_RECORDS):
        employee_id = f"BP{index + 1:04d}"
        first_name = first_names[index]
        last_name = rng.choice(LAST_NAMES)
        department = weighted_choice(rng, dept_weights)
        position = rng.choice(POSITIONS_BY_DEPT[department])
        province = weighted_choice(rng, province_weights)
        city = rng.choice(CITIES_BY_PROVINCE[province])

        if rng.random() < 0.76:
            start = recent_start
            span = (recent_end - recent_start).days
        else:
            start = older_start
            span = (older_end - older_start).days
        hire_date = start + timedelta(days=rng.randint(0, span))

        salary_low, salary_high = POSITION_SALARY[position]
        salary = rng.randrange(salary_low, salary_high + 1, 500)
        email = f"{first_name.lower()}.{last_name.lower().replace(' ', '')}@brightpath.com"
        phone = f"09{rng.randint(10_000_000, 99_999_999):08d}"
        status = weighted_choice(rng, STATUS_WEIGHTS)

        records.append(
            {
                "EmployeeID": employee_id,
                "FirstName": first_name,
                "LastName": last_name,
                "Email": email,
                "Phone": phone,
                "Department": department,
                "Position": position,
                "City": city,
                "Province": province,
                "HireDate": hire_date,
                "Salary": salary,
                "EmploymentStatus": status,
            }
        )

    return records


def style_headers(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_clean_workbook(records: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Clean Employee Data"

    headers = list(records[0].keys())
    ws.append(headers)
    for record in records:
        ws.append([record[column] for column in headers])

    for row in ws.iter_rows(min_row=2, min_col=10, max_col=10):
        for cell in row:
            cell.number_format = "yyyy-mm-dd"
    for row in ws.iter_rows(min_row=2, min_col=11, max_col=11):
        for cell in row:
            cell.number_format = '#,##0'

    style_headers(ws)
    widths = {"A": 14, "B": 14, "C": 16, "D": 30, "E": 15, "F": 24, "G": 31, "H": 17, "I": 14, "J": 14, "K": 14, "L": 18}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    wb.save(CLEAN_FILE)


def format_date_messy(rng: random.Random, value: date) -> str:
    return rng.choice(
        [
            value.strftime("%Y-%m-%d"),
            value.strftime("%m/%d/%Y"),
            value.strftime("%B %d, %Y"),
            value.strftime("%d-%b-%y"),
        ]
    )


def add_padding(rng: random.Random, value: str) -> str:
    left = " " * rng.randint(1, 2) if rng.random() < 0.5 else ""
    right = " " * rng.randint(1, 2) if rng.random() < 0.5 else ""
    return f"{left}{value}{right}"


def case_variant(rng: random.Random, value: str) -> str:
    return rng.choice([value, value.lower(), value.upper(), value.title()])


def build_raw_records(records: list[dict[str, object]]) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    rng = random.Random(SEED + 1)
    raw = [dict(record) for record in records]

    issue_counts: Counter[str] = Counter()

    # Duplicate records: 120 rows, some exact and some with tiny formatting drift.
    duplicate_source_indices = rng.sample(range(len(raw)), 120)
    duplicate_rows: list[dict[str, object]] = []
    for index in duplicate_source_indices:
        row = dict(raw[index])
        if rng.random() < 0.45:
            row["FirstName"] = case_variant(rng, str(row["FirstName"]))
        if rng.random() < 0.45:
            row["LastName"] = case_variant(rng, str(row["LastName"]))
        if rng.random() < 0.35:
            row["Department"] = add_padding(rng, str(row["Department"]))
        duplicate_rows.append(row)

    insertion_points = sorted(rng.sample(range(len(raw) + len(duplicate_rows)), len(duplicate_rows)))
    merged: list[dict[str, object]] = []
    raw_iter = iter(raw)
    dup_iter = iter(duplicate_rows)
    insert_set = set(insertion_points)
    for slot in range(len(raw) + len(duplicate_rows)):
        if slot in insert_set:
            merged.append(next(dup_iter))
        else:
            merged.append(next(raw_iter))
    raw = merged
    issue_counts["Duplicate Records"] = 120

    name_rows = rng.sample(range(len(raw)), 300)
    dept_rows = rng.sample(range(len(raw)), 250)
    space_rows = rng.sample(range(len(raw)), 300)
    date_rows = rng.sample(range(len(raw)), 400)
    missing_rows = rng.sample(range(len(raw)), 150)
    email_rows = rng.sample(range(len(raw)), 60)
    phone_rows = rng.sample(range(len(raw)), 250)
    salary_rows = rng.sample(range(len(raw)), 200)
    status_rows = rng.sample(range(len(raw)), 100)

    for row_index in name_rows:
        row = raw[row_index]
        row["FirstName"] = case_variant(rng, str(row["FirstName"]))
        row["LastName"] = case_variant(rng, str(row["LastName"]))
        if rng.random() < 0.25:
            row["FirstName"] = add_padding(rng, str(row["FirstName"]))
        issue_counts["Name Formatting Issues"] += 1

    department_variants = {
        "Human Resources": ["HR", "human resources", "Human Resource", "H.R."],
        "Information Technology": ["IT", "information technology", "Info Tech"],
        "Customer Service": ["Customer service", "customer support"],
        "Operations": ["operations", "Ops"],
        "Finance": ["finance"],
        "Marketing": ["Marketing ", " marketing"],
    }
    for row_index in dept_rows:
        row = raw[row_index]
        current = str(row["Department"])
        row["Department"] = rng.choice(department_variants.get(current, [current.lower(), current.upper()]))
        issue_counts["Department Inconsistencies"] += 1

    for row_index in space_rows:
        row = raw[row_index]
        column = rng.choice(["Department", "Position", "Email", "Phone", "Salary", "EmploymentStatus"])
        row[column] = add_padding(rng, str(row[column]))
        issue_counts["Extra Spaces"] += 1

    for row_index in date_rows:
        raw[row_index]["HireDate"] = format_date_messy(rng, raw[row_index]["HireDate"])
        issue_counts["Date Format Problems"] += 1

    for row_index in missing_rows:
        column = rng.choice(["Email", "Phone", "Department", "Position"])
        raw[row_index][column] = ""
        issue_counts["Missing Values"] += 1

    for row_index in email_rows:
        row = raw[row_index]
        base = str(row["Email"]).split("@")[0]
        row["Email"] = rng.choice([
            f"{base}.gmail.com",
            f"{base}@@gmail.com",
            "@gmail.com",
            f"{base}email.com",
            f"{base}@brightpathcom",
        ])
        issue_counts["Email Errors"] += 1

    for row_index in phone_rows:
        row = raw[row_index]
        digits = str(row["Phone"]).replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
        row["Phone"] = rng.choice([
            digits,
            digits[1:],
            f"+63{digits[1:]}",
            f"{digits[:4]}-{digits[4:7]}-{digits[7:]}",
            f"({digits[:4]}) {digits[4:7]}-{digits[7:]}",
        ])
        issue_counts["Phone Number Formatting Problems"] += 1

    for row_index in salary_rows:
        salary_text = str(raw[row_index]["Salary"]).replace(",", "")
        salary_value = int(float(salary_text))
        raw[row_index]["Salary"] = rng.choice([f"{salary_value}", f"{salary_value:,}", f"₱{salary_value:,}", f"{salary_value:.2f}"])
        issue_counts["Salary Formatting Issues"] += 1

    for row_index in status_rows:
        current = str(raw[row_index]["EmploymentStatus"])
        raw[row_index]["EmploymentStatus"] = rng.choice([current, current.upper(), current.lower(), "Actve" if current == "Active" else current])
        issue_counts["Employment Status Issues"] += 1

    issue_rows = [
        {"Issue Type": "Duplicate Records", "Number of Affected Records": issue_counts["Duplicate Records"], "Example": "Repeated Employee IDs and near-duplicate rows"},
        {"Issue Type": "Name Formatting Issues", "Number of Affected Records": issue_counts["Name Formatting Issues"], "Example": "maria santos / MARIA SANTOS / Maria santos"},
        {"Issue Type": "Department Inconsistencies", "Number of Affected Records": issue_counts["Department Inconsistencies"], "Example": "HR / H.R. / Info Tech / customer support"},
        {"Issue Type": "Extra Spaces", "Number of Affected Records": issue_counts["Extra Spaces"], "Example": " Finance  /  Marketing / Sales "},
        {"Issue Type": "Date Format Problems", "Number of Affected Records": issue_counts["Date Format Problems"], "Example": "05/10/2024 / May 10, 2024 / 10-May-24"},
        {"Issue Type": "Missing Values", "Number of Affected Records": issue_counts["Missing Values"], "Example": "Blank email, phone, department, or position cells"},
        {"Issue Type": "Email Errors", "Number of Affected Records": issue_counts["Email Errors"], "Example": "john.gmail.com / john@@gmail.com / @gmail.com"},
        {"Issue Type": "Phone Number Formatting Problems", "Number of Affected Records": issue_counts["Phone Number Formatting Problems"], "Example": "+639171234567 / 0917-123-4567 / 9171234567"},
        {"Issue Type": "Salary Formatting Issues", "Number of Affected Records": issue_counts["Salary Formatting Issues"], "Example": "35000 / 35,000 / ₱35,000 / 35000.00"},
        {"Issue Type": "Employment Status Issues", "Number of Affected Records": issue_counts["Employment Status Issues"], "Example": "ACTIVE / active / Actve"},
    ]

    return raw, issue_rows


def write_raw_workbook(raw_records: list[dict[str, object]], issue_rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Employee Data"
    headers = list(raw_records[0].keys())
    ws.append(headers)
    for record in raw_records:
        ws.append([record[column] for column in headers])

    style_headers(ws)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 31
    ws.column_dimensions["H"].width = 17
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 18
    ws.column_dimensions["K"].width = 14
    ws.column_dimensions["L"].width = 18

    issues_ws = wb.create_sheet("Data Quality Issues")
    issue_headers = list(issue_rows[0].keys())
    issues_ws.append(issue_headers)
    for row in issue_rows:
        issues_ws.append([row[column] for column in issue_headers])
    style_headers(issues_ws)
    issues_ws.column_dimensions["A"].width = 34
    issues_ws.column_dimensions["B"].width = 26
    issues_ws.column_dimensions["C"].width = 48

    wb.save(RAW_FILE)


def write_readme() -> None:
    README_FILE.write_text(
        """# Employee Records Cleanup & Validation for Business Reporting

**Project Name:** Employee Records Cleanup & Validation for Business Reporting  
**Client:** BrightPath Solutions Inc. (Fictional)  
**Role:** Data Entry and Data Processing Specialist  
**Dataset Size:** 3,500 records  
**Purpose:** Clean and prepare employee records for HR system migration.

## Project Overview

This case study simulates an HR employee database received from BrightPath Solutions Inc., a fictional BPO company based in the Philippines. The file set demonstrates how inconsistent manual data entry can affect reporting, validation, and migration workflows.

## Dataset Description

The dataset contains 3,500 employee records with the following fields: EmployeeID, FirstName, LastName, Email, Phone, Department, Position, City, Province, HireDate, Salary, and EmploymentStatus. The clean workbook represents the maintained source of truth, while the raw workbook introduces realistic issues for a cleaning exercise.

## Data Quality Issues Identified

- Duplicate records and repeated EmployeeIDs
- Inconsistent capitalization and spacing in names
- Department abbreviations and non-standard labels
- Mixed date formats across hire dates
- Missing values in key HR fields
- Invalid email formats
- Phone number formatting inconsistencies
- Salary formatting differences
- Employment status variations and typos

## Cleaning Tasks Performed

- Standardized names, departments, and employment status values
- Removed duplicate rows
- Validated email and phone number formats
- Normalized hire dates into one date format
- Checked salary consistency and formatting
- Preserved complete employee IDs for traceability
- Organized the final dataset for reporting and migration use

## Tools Used

- Microsoft Excel
- Data validation and spreadsheet formatting
- Python for controlled dataset generation

## Expected Results

The clean workbook is ready for HR reporting and system migration, while the raw workbook provides a realistic case study for demonstrating data cleaning, validation, and spreadsheet organization skills.
""",
        encoding="utf-8",
    )


def validate_outputs() -> None:
    clean_wb = load_workbook(CLEAN_FILE, data_only=True)
    raw_wb = load_workbook(RAW_FILE, data_only=True)

    clean_ws = clean_wb["Clean Employee Data"]
    raw_ws = raw_wb["Raw Employee Data"]
    issue_ws = raw_wb["Data Quality Issues"]

    clean_rows = list(clean_ws.iter_rows(min_row=2, values_only=True))
    raw_rows = list(raw_ws.iter_rows(min_row=2, values_only=True))

    assert clean_wb.sheetnames == ["Clean Employee Data"]
    assert raw_wb.sheetnames == ["Raw Employee Data", "Data Quality Issues"]
    assert len(clean_rows) == TOTAL_RECORDS
    assert len({row[0] for row in clean_rows}) == TOTAL_RECORDS
    assert sum(1 for row in clean_rows for value in row if value in (None, "")) == 0
    assert len(raw_rows) >= TOTAL_RECORDS
    assert issue_ws.max_row == 11

    print(f"Validated clean rows: {len(clean_rows)}")
    print(f"Validated raw rows: {len(raw_rows)}")
    print(f"Validated issue summary rows: {issue_ws.max_row - 1}")


def validate_clean(records: list[dict[str, object]]) -> None:
    assert len(records) == TOTAL_RECORDS
    assert len({record["EmployeeID"] for record in records}) == TOTAL_RECORDS
    for record in records:
        for value in record.values():
            assert value not in (None, "")


def main() -> None:
    records = generate_clean_records()
    validate_clean(records)
    write_clean_workbook(records)
    raw_records, issue_rows = build_raw_records(records)
    write_raw_workbook(raw_records, issue_rows)
    write_readme()
    validate_outputs()
    print(f"Created {CLEAN_FILE}")
    print(f"Created {RAW_FILE}")
    print(f"Created {README_FILE}")


if __name__ == "__main__":
    main()